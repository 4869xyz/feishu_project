"""End-to-end tests for signing-only XLSX aggregation."""

from __future__ import annotations

from copy import copy
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, Font, PatternFill
from openpyxl.writer.theme import theme_xml

from services.sales_workbook_aggregator import (
    DuplicateSourceError,
    SalesAggregationError,
    SIGNING_ACCOUNTING_FORMAT,
    SIGNING_AMOUNT_MIN_WIDTH,
    SIGNING_AMOUNT_WIDTH_PADDING,
    SourceValidationError,
    SourceWorkbook,
    aggregate_sales_workbooks,
    validate_source_workbook,
)


SIGNING_TARGET = "展示用-2026年签约数据汇总（0702统计）"
REPAYMENT_TARGET = "展示用-2026年回款数据汇总(0702统计)"


def _write_signing_headers(sheet) -> None:
    headers = [
        "组别",
        "序号",
        "人员",
        "客户简称",
        "销售产品名称/项目名称",
        "落地地点",
        "客户来源\n（展会需注明）",
        "备注",
        "签约后且收到第一笔款后才算已签约",
    ]
    for column, value in enumerate(headers, 1):
        sheet.cell(2, column).value = value
    for index in range(12):
        sheet.cell(3, 9 + index).value = f"{index + 1}月"


def _write_signing_rows(sheet, rows: list[list]) -> None:
    _write_signing_headers(sheet)
    for row, values in enumerate(rows, 4):
        for column, value in enumerate(values, 1):
            sheet.cell(row, column).value = value


def _make_source(
    path: Path,
    signing_rows: list[list],
    *,
    signing_sheet_name: str = "签约数据",
    other_sheet_names: tuple[str, ...] = (),
) -> None:
    workbook = Workbook()
    signing = workbook.active
    signing.title = signing_sheet_name
    _write_signing_rows(signing, signing_rows)
    for name in other_sheet_names:
        other = workbook.create_sheet(name)
        other["A1"] = "该工作表应被忽略"
        other["F4"] = "回款金额故意写成非法文本"
    workbook.save(path)


def _source_row(
    *,
    group=None,
    sequence=None,
    person=None,
    customer=None,
    product=None,
    location=None,
    customer_source=None,
    note=None,
    months=(),
) -> list:
    """Return one A:T source row with selected test values."""

    values = [None] * 20
    values[0] = group
    values[1] = sequence
    values[2] = person
    values[3] = customer
    values[4] = product
    values[5] = location
    values[6] = customer_source
    values[7] = note
    for index, value in enumerate(months):
        values[8 + index] = value
    return values


def _make_candidate_source(path: Path, sheet_names: list[str]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, name in enumerate(sheet_names, 1):
        sheet = workbook.create_sheet(name)
        _write_signing_rows(
            sheet,
            [[f"第{index}组", 1, f"人员{index}", None, None, None, None, None, index]],
        )
    workbook.save(path)


def _style_row(sheet, row: int, max_column: int, color: str) -> None:
    for column in range(1, max_column + 1):
        cell = sheet.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name="Microsoft YaHei", bold=row != 4)
    sheet.row_dimensions[row].height = 20 + row % 3


def _font_color_signature(cell) -> tuple | None:
    color = cell.font.color
    if color is None:
        return None
    if color.type == "rgb":
        value = color.rgb
    elif color.type == "theme":
        value = color.theme
    elif color.type == "indexed":
        value = color.indexed
    elif color.type == "auto":
        value = color.auto
    else:
        value = None
    return color.type, value, color.tint


def _style_without_border(cell) -> tuple:
    return (
        copy(cell.font),
        copy(cell.fill),
        copy(cell.alignment),
        cell.number_format,
        copy(cell.protection),
    )


def _is_thin_black_side(side) -> bool:
    return (
        side.style == "thin"
        and side.color is not None
        and side.color.type == "rgb"
        and side.color.rgb == "FF000000"
    )


def _has_any_box_side(cell) -> bool:
    return any(
        side is not None and side.style is not None
        for side in (
            cell.border.left,
            cell.border.right,
            cell.border.top,
            cell.border.bottom,
        )
    )


def _make_template(path: Path, *, include_repayment: bool = True) -> None:
    workbook = Workbook()
    signing = workbook.active
    signing.title = SIGNING_TARGET
    _write_signing_headers(signing)
    signing["A4"] = "样例组"
    signing["B4"] = 1
    signing["C4"] = "样例人员"
    labels = {
        5: "个人月度合计：",
        6: "个人季度合计：",
        7: "个人年度合计：",
        9: "小组月度合计：",
        10: "小组季度合计：",
        11: "小组年度合计：",
    }
    for row, label in labels.items():
        signing.cell(row, 2).value = label
    signing["A13"] = "部门月度合计："
    signing["A14"] = "部门季度合计："
    signing["A15"] = "部门年度合计："
    for row in range(4, 16):
        _style_row(signing, row, 20, f"{row:02X}{row:02X}EE")
    signing.row_dimensions[12].hidden = True

    ranking = workbook.create_sheet("展示用-签约排名汇总（0702统计）")
    ranking["A1"] = "必须保留"
    ranking["B2"] = "=1+1"

    if include_repayment:
        repayment = workbook.create_sheet(REPAYMENT_TARGET)
        repayment["A1"] = "回款表必须原样保留"
        repayment["F4"] = "旧回款内容"
        repayment["G5"] = "=1+1"
        repayment.row_dimensions[6].hidden = True

    detail = workbook.create_sheet("明细")
    detail["A1"] = "必须保留"
    workbook.save(path)


@pytest.mark.parametrize("sheet_name", ["Sheet1", "签约数据", "任意名称"])
def test_single_sheet_uses_its_only_worksheet(
    project_tmp_dir: Path, sheet_name: str
) -> None:
    source = project_tmp_dir / f"single-{sheet_name}.xlsx"
    _make_source(
        source,
        [["甲组", 1, "A", None, None, None, None, None, 10]],
        signing_sheet_name=sheet_name,
    )

    result = validate_source_workbook(SourceWorkbook("single", source))

    assert result.signing_sheet_name == sheet_name
    assert result.signing_detail_count == 1


@pytest.mark.parametrize(
    ("sheet_names", "expected"),
    [
        (["销售签约数据", "签约数据", "签约情况"], "签约情况"),
        (["销售签约数据", "签约数据"], "签约数据"),
        (["销售签约数据", "2026签约情况"], "2026签约情况"),
        (["销售签约数据", "历史签约数据"], "销售签约数据"),
    ],
)
def test_multi_sheet_uses_confirmed_name_priority(
    project_tmp_dir: Path, sheet_names: list[str], expected: str
) -> None:
    source = project_tmp_dir / "candidates.xlsx"
    _make_candidate_source(source, sheet_names)

    result = validate_source_workbook(SourceWorkbook("candidate", source))

    assert result.signing_sheet_name == expected
    assert result.signing_detail_count == 1


def test_multi_sheet_without_candidate_lists_available_names(
    project_tmp_dir: Path,
) -> None:
    source = project_tmp_dir / "no-signing-source.xlsx"
    _make_candidate_source(source, ["签约汇总", "签约排名"])

    with pytest.raises(SourceValidationError) as exc_info:
        validate_source_workbook(SourceWorkbook("missing", source))

    message = str(exc_info.value)
    assert "未找到" in message
    assert "签约汇总" in message
    assert "签约排名" in message


def test_aggregate_rebuilds_signing_and_preserves_repayment(
    project_tmp_dir: Path,
) -> None:
    template = project_tmp_dir / "template.xlsx"
    source1 = project_tmp_dir / "source1.xlsx"
    source2 = project_tmp_dir / "source2.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    _make_source(
        source1,
        [
            ["甲组", 1, "A", "客户1", "项目", None, None, None, 10],
            [None, 2, "A", "客户2", "项目", None, None, None, None, 20],
        ],
        other_sheet_names=("回款数据",),
    )
    _make_source(
        source2,
        [
            ["甲组", 1, "A", "客户1", "项目", None, None, None, 10],
            ["乙组", 1, "B", "客户3", "项目", None, None, None, None, None, 5],
        ],
        signing_sheet_name="2026签约情况",
        other_sheet_names=("损坏的回款情况",),
    )

    result = aggregate_sales_workbooks(
        [SourceWorkbook("one", source1), SourceWorkbook("two", source2)],
        template,
        output,
    )

    assert result.signing_detail_count == 4
    assert result.signing_total == Decimal(45)

    workbook = load_workbook(output, data_only=False)
    signing = workbook[SIGNING_TARGET]
    assert signing.max_row == 26
    assert signing["A4"].value == "甲组"
    assert signing["A15"].value == "乙组"
    assert signing["I7"].value == "=SUM(I4:I6)"
    assert signing["I11"].value == "=SUM(I7)"
    assert signing["I24"].value == "=SUM(I11,I20)"
    assert "A4:A13" in {str(item) for item in signing.merged_cells.ranges}
    assert "A15:A22" in {str(item) for item in signing.merged_cells.ranges}
    assert all(not signing.row_dimensions[row].hidden for row in range(4, 27))
    assert all(signing.cell(10, column).value is None for column in range(1, 21))
    assert all(signing.cell(14, column).value is None for column in range(1, 21))
    assert all(signing.cell(19, column).value is None for column in range(1, 21))
    assert all(signing.cell(23, column).value is None for column in range(1, 21))
    for coordinate in ("A1", "T1", "A3", "T3", "T24"):
        border = signing[coordinate].border
        assert all(
            _is_thin_black_side(side)
            for side in (border.left, border.right, border.top, border.bottom)
        )
    assert _is_thin_black_side(signing["A4"].border.top)
    assert _is_thin_black_side(signing["A13"].border.bottom)
    assert all(
        _is_thin_black_side(side)
        for side in (
            signing["A10"].border.left,
            signing["A10"].border.right,
            signing["A19"].border.left,
            signing["A19"].border.right,
        )
    )
    for coordinate in ("A14", "A23"):
        border = signing[coordinate].border
        assert all(
            _is_thin_black_side(side)
            for side in (border.left, border.right, border.top, border.bottom)
        )
    for spacer_row in (10, 14, 19, 23):
        assert all(
            not _has_any_box_side(signing.cell(spacer_row, column))
            for column in range(2, 21)
        )
    assert all(
        _is_thin_black_side(side)
        for side in (
            signing["I26"].border.left,
            signing["I26"].border.top,
            signing["I26"].border.bottom,
            signing["T26"].border.right,
            signing["T26"].border.top,
            signing["T26"].border.bottom,
        )
    )
    for column in range(1, 21):
        border = signing.cell(27, column).border
        assert not any(
            _is_thin_black_side(side)
            for side in (border.left, border.right, border.top, border.bottom)
        )

    template_workbook = load_workbook(template, data_only=False)
    template_signing = template_workbook[SIGNING_TARGET]
    assert _style_without_border(signing["B10"]) == _style_without_border(
        template_signing["B8"]
    )
    assert signing.row_dimensions[10].height == template_signing.row_dimensions[8].height
    assert _style_without_border(signing["B23"]) == _style_without_border(
        template_signing["B12"]
    )
    assert signing.row_dimensions[23].height == template_signing.row_dimensions[12].height
    template_workbook.close()

    repayment = workbook[REPAYMENT_TARGET]
    assert repayment["A1"].value == "回款表必须原样保留"
    assert repayment["F4"].value == "旧回款内容"
    assert repayment["G5"].value == "=1+1"
    assert repayment.row_dimensions[6].hidden is True
    assert workbook["展示用-签约排名汇总（0702统计）"]["A1"].value == "必须保留"
    assert workbook["明细"]["A1"].value == "必须保留"
    workbook.close()


def test_template_without_repayment_sheet_still_succeeds(project_tmp_dir: Path) -> None:
    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "source.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template, include_repayment=False)
    _make_source(
        source,
        [["甲组", 1, "A", None, None, None, None, None, 1]],
        signing_sheet_name="任意单表名称",
    )

    aggregate_sales_workbooks([SourceWorkbook("one", source)], template, output)

    workbook = load_workbook(output, read_only=True)
    assert SIGNING_TARGET in workbook.sheetnames
    assert REPAYMENT_TARGET not in workbook.sheetnames
    workbook.close()


def test_duplicate_source_id_is_rejected(project_tmp_dir: Path) -> None:
    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "source.xlsx"
    _make_template(template)
    _make_source(
        source,
        [["甲组", 1, "A", None, None, None, None, None, 1]],
        signing_sheet_name="Sheet1",
    )
    item = SourceWorkbook("same", source)
    with pytest.raises(DuplicateSourceError):
        aggregate_sales_workbooks([item, item], template, project_tmp_dir / "out.xlsx")


def test_invalid_amount_reports_selected_sheet_row_and_column(
    project_tmp_dir: Path,
) -> None:
    source = project_tmp_dir / "source.xlsx"
    _make_source(
        source,
        [["甲组", 1, "A", None, None, None, None, None, "不是金额"]],
        signing_sheet_name="任意名称",
    )

    with pytest.raises(SourceValidationError, match="任意名称.*第 4 行.*I 列"):
        validate_source_workbook(SourceWorkbook("source", source))


def test_blank_person_summary_and_note_rows_are_ignored(
    project_tmp_dir: Path,
) -> None:
    """Summary values and annotations never enter signing validation or totals."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "source-with-summary.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    rows = [
        _source_row(group="甲组", sequence=1, person="A", months=(10,)),
        _source_row(sequence="个人月度合计：", months=(10, " - ")),
        _source_row(sequence="个人季度合计：", months=(10,)),
        _source_row(sequence="个人年度合计：", months=(10,)),
        _source_row(sequence="有修改的请用别的颜色填充", note="说明", months=(" - ",)),
        _source_row(sequence=None, person="A", months=(None, 5)),
        _source_row(sequence=99, person=None, note="漏填人员", months=(999,)),
    ]
    _make_source(source, rows, signing_sheet_name="签约情况")
    workbook = load_workbook(source)
    sheet = workbook["签约情况"]
    for row in range(5, 8):
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    workbook.save(source)
    workbook.close()

    validation = validate_source_workbook(SourceWorkbook("source", source))
    result = aggregate_sales_workbooks(
        [SourceWorkbook("source", source)], template, output
    )

    assert validation.signing_detail_count == 2
    assert result.signing_detail_count == 2
    assert result.signing_total == Decimal(15)


def test_all_summary_labels_do_not_replace_inherited_group(
    project_tmp_dir: Path,
) -> None:
    """All nine summary labels are skipped before they can alter group inheritance."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "all-summary-labels.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    labels = (
        "个人月度合计",
        "个人季度合计：",
        "个人年度合计",
        "小组月度合计：",
        "小组季度合计",
        "小组年度合计：",
        "部门月度合计",
        "部门季度合计：",
        "部门年度合计",
    )
    rows = [_source_row(group="甲组", sequence=1, person="A", months=(1,))]
    for label in labels[:6]:
        rows.append(_source_row(sequence=label, months=(" - ",)))
    for label in labels[6:]:
        rows.append(_source_row(group=label, months=(" - ",)))
    rows.append(_source_row(sequence=2, person="B", months=(2,)))
    _make_source(source, rows, signing_sheet_name="签约数据")

    result = aggregate_sales_workbooks(
        [SourceWorkbook("source", source)], template, output
    )

    assert result.signing_detail_count == 2
    assert result.signing_total == Decimal(3)
    workbook = load_workbook(output, data_only=False)
    signing = workbook[SIGNING_TARGET]
    assert signing["A4"].value == "甲组"
    vertical_group_merges = [
        merged
        for merged in signing.merged_cells.ranges
        if merged.min_col == merged.max_col == 1 and merged.min_row >= 4
    ]
    assert len(vertical_group_merges) == 1
    workbook.close()


def test_group_only_row_sets_group_for_following_detail(project_tmp_dir: Path) -> None:
    """An A-only group header remains meaningful even though its person is blank."""

    source = project_tmp_dir / "group-header.xlsx"
    _make_source(
        source,
        [
            _source_row(group="甲组"),
            _source_row(sequence=1, person="A", months=(1,)),
        ],
        signing_sheet_name="Sheet1",
    )

    validation = validate_source_workbook(SourceWorkbook("source", source))

    assert validation.signing_detail_count == 1


@pytest.mark.parametrize("person", [123, True])
def test_non_text_non_blank_person_is_rejected(
    project_tmp_dir: Path,
    person,
) -> None:
    """Only blank personnel cells are ignored; malformed nonblank values still fail."""

    source = project_tmp_dir / f"invalid-person-{person}.xlsx"
    _make_source(
        source,
        [_source_row(group="甲组", sequence=1, person=person, months=(1,))],
        signing_sheet_name="Sheet1",
    )

    with pytest.raises(SourceValidationError, match="第 4 行.*C 列.*人员必须是文本"):
        validate_source_workbook(SourceWorkbook("source", source))


def test_detail_cells_preserve_source_font_colors_only(
    project_tmp_dir: Path,
) -> None:
    """Each detail cell keeps its source color while all other styles stay templated."""

    template = project_tmp_dir / "template.xlsx"
    source1 = project_tmp_dir / "source-red.xlsx"
    source2 = project_tmp_dir / "source-black.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    _make_source(
        source1,
        [
            _source_row(group="甲组", sequence=1, person="同一人", months=(10,)),
            _source_row(sequence="个人月度合计：", months=(10,)),
        ],
        signing_sheet_name="签约数据",
    )
    _make_source(
        source2,
        [_source_row(group="甲组", sequence=2, person="同一人", months=(20,))],
        signing_sheet_name="签约数据",
    )

    workbook = load_workbook(template)
    sheet = workbook[SIGNING_TARGET]
    for column in range(1, 21):
        sheet.cell(4, column).font = Font(
            name="Microsoft YaHei",
            size=11,
            bold=False,
            italic=False,
            color=Color(rgb="FF7030A0"),
        )
        sheet.cell(5, column).font = Font(
            name="Microsoft YaHei", size=12, bold=True, color=Color(rgb="FF0070C0")
        )
    sheet["I4"].number_format = "#,##0"
    workbook.save(template)
    workbook.close()

    workbook = load_workbook(source1)
    sheet = workbook["签约数据"]
    sheet["A4"].font = Font(
        name="Arial", size=22, bold=True, color=Color(rgb="FFC00000")
    )
    sheet["B4"].font = Font(
        name="Arial", size=22, bold=True, color=Color(theme=4, tint=0.25)
    )
    sheet["C4"].font = Font(
        name="Arial", size=22, bold=True, color=Color(rgb="FFFF0000")
    )
    sheet["D4"].font = Font(name="Arial", size=22, bold=True, color=Color(indexed=8))
    sheet["E4"].font = Font(name="Arial", size=22, bold=True, color=Color(auto=True))
    sheet["F4"].font = Font(name="Arial", size=22, bold=True)
    sheet["I4"].font = Font(
        name="Arial", size=22, bold=True, color=Color(rgb="FF373C43")
    )
    sheet["I4"].number_format = "0.00"
    sheet["B5"].font = Font(color=Color(rgb="FF00B050"))
    workbook.save(source1)
    workbook.close()

    workbook = load_workbook(source2)
    sheet = workbook["签约数据"]
    sheet["C4"].font = Font(
        name="SimSun", size=18, bold=True, color=Color(rgb="FF000000")
    )
    workbook.save(source2)
    workbook.close()

    aggregate_sales_workbooks(
        [SourceWorkbook("red", source1), SourceWorkbook("black", source2)],
        template,
        output,
    )

    output_workbook = load_workbook(output, data_only=False)
    output_sheet = output_workbook[SIGNING_TARGET]
    template_workbook = load_workbook(template, data_only=False)
    template_sheet = template_workbook[SIGNING_TARGET]

    assert _font_color_signature(output_sheet["A4"]) == ("rgb", "FFC00000", 0.0)
    assert _font_color_signature(output_sheet["B4"]) == ("theme", 4, 0.25)
    assert _font_color_signature(output_sheet["C4"]) == ("rgb", "FFFF0000", 0.0)
    assert _font_color_signature(output_sheet["D4"]) == ("indexed", 8, 0.0)
    assert _font_color_signature(output_sheet["E4"]) == ("auto", True, 0.0)
    assert _font_color_signature(output_sheet["F4"]) is None
    assert _font_color_signature(output_sheet["I4"]) == ("rgb", "FF373C43", 0.0)
    assert _font_color_signature(output_sheet["C5"]) == ("rgb", "FF000000", 0.0)

    assert output_sheet["C4"].font.name == template_sheet["C4"].font.name
    assert output_sheet["C4"].font.sz == template_sheet["C4"].font.sz
    assert output_sheet["C4"].font.bold == template_sheet["C4"].font.bold
    assert output_sheet["C4"].font.italic == template_sheet["C4"].font.italic
    assert output_sheet["C4"].fill.fill_type == template_sheet["C4"].fill.fill_type
    assert output_sheet["C4"].fill.fgColor.rgb == template_sheet["C4"].fill.fgColor.rgb
    assert (
        output_sheet["C4"].alignment.horizontal
        == template_sheet["C4"].alignment.horizontal
    )
    assert output_sheet["C4"].border.left.style == "thin"
    assert output_sheet["C4"].border.left.color.rgb == "FF000000"
    assert output_sheet["I4"].number_format == SIGNING_ACCOUNTING_FORMAT
    assert output_sheet["I4"].number_format != template_sheet["I4"].number_format

    # The generated personal total uses the template sample, never the ignored source total.
    assert _font_color_signature(output_sheet["B6"]) == ("rgb", "FF0070C0", 0.0)
    output_workbook.close()
    template_workbook.close()


def test_signing_amounts_use_symbol_free_accounting_with_one_decimal(
    project_tmp_dir: Path,
) -> None:
    """Detail and summary amounts share one accounting format without changing values."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "accounting-source.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    _make_source(
        source,
        [
            _source_row(
                group="甲组",
                sequence=1,
                person="同一人",
                note="格式验证",
                months=(1234, 0, -12.34, None),
            )
        ],
        signing_sheet_name="签约数据",
    )
    workbook = load_workbook(source)
    sheet = workbook["签约数据"]
    sheet["C4"].font = Font(color=Color(rgb="FFED7D31"))
    sheet["I4"].number_format = "0.000"
    workbook.save(source)
    workbook.close()

    result = aggregate_sales_workbooks(
        [SourceWorkbook("accounting", source)], template, output
    )

    assert result.signing_total == Decimal("1221.66")
    workbook = load_workbook(output, data_only=False)
    sheet = workbook[SIGNING_TARGET]
    assert (sheet["I4"].value, sheet["J4"].value, sheet["K4"].value, sheet["L4"].value) == (
        1234,
        0,
        -12.34,
        None,
    )
    assert sheet.row_dimensions[4].hidden is True
    assert all(
        sheet.cell(4, column).number_format == SIGNING_ACCOUNTING_FORMAT
        for column in range(9, 21)
    )
    for row in range(4, sheet.max_row + 1):
        for column in range(9, 21):
            cell = sheet.cell(row, column)
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            if isinstance(cell.value, (int, float)) or is_formula:
                assert cell.number_format == SIGNING_ACCOUNTING_FORMAT
    assert sheet["I5"].value == "=SUM(I4:I4)"
    assert sheet["I6"].value == "=SUM(I5:K5)"
    assert sheet["I7"].value == "=SUM(I5:T5)"
    assert sheet["I8"].value is None
    assert sheet["I8"].number_format != SIGNING_ACCOUNTING_FORMAT
    assert "$" not in SIGNING_ACCOUNTING_FORMAT
    assert "¥" not in SIGNING_ACCOUNTING_FORMAT
    assert "0.0" in SIGNING_ACCOUNTING_FORMAT
    workbook.close()


def test_signing_amount_columns_expand_independently_for_displayed_values(
    project_tmp_dir: Path,
) -> None:
    """I:T retain a template floor and expand per month for large accounting values."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "wide-amounts.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    workbook = load_workbook(template)
    sheet = workbook[SIGNING_TARGET]
    sheet.column_dimensions["I"].width = 8
    sheet.column_dimensions["J"].width = 9
    sheet.column_dimensions["K"].width = 25
    workbook.save(template)
    workbook.close()
    _make_source(
        source,
        [
            _source_row(
                group="甲组",
                sequence=1,
                person="A",
                months=(1, 1234567890.1, -9876543.2),
            ),
            _source_row(
                sequence=2,
                person="A",
                months=(2, 2000000000.2, 0),
            ),
        ],
        signing_sheet_name="签约数据",
    )

    aggregate_sales_workbooks(
        [SourceWorkbook("wide-amounts", source)], template, output
    )

    workbook = load_workbook(output, data_only=False)
    sheet = workbook[SIGNING_TARGET]
    expected_j = float(len(f"{Decimal('3234567890.3'):,.1f}") + SIGNING_AMOUNT_WIDTH_PADDING)
    assert sheet.column_dimensions["I"].width == SIGNING_AMOUNT_MIN_WIDTH
    assert sheet.column_dimensions["J"].width == expected_j
    assert sheet.column_dimensions["J"].width > sheet.column_dimensions["I"].width
    assert sheet.column_dimensions["K"].width == 25
    for column in "IJKLMNOPQRST":
        dimension = sheet.column_dimensions[column]
        assert dimension.bestFit is True
        assert dimension.width >= SIGNING_AMOUNT_MIN_WIDTH
    workbook.close()


def test_yellow_orange_key_fields_hide_rows_without_losing_data_or_totals(
    project_tmp_dir: Path,
) -> None:
    """Warm key fields hide rows while month-only yellow and summaries stay visible."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "yellow-rows.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    rows = [
        _source_row(
            group="甲组" if index == 1 else None,
            sequence=index,
            person="同一人",
            customer_source="展会" if index == 13 else None,
            note=f"明细{index}",
            months=(index,),
        )
        for index in range(1, 14)
    ]
    _make_source(source, rows, signing_sheet_name="签约数据")

    workbook = load_workbook(source)
    sheet = workbook["签约数据"]
    solid_rows = {
        4: "FFFFC000",
        5: "FFFFC60A",
        6: "FFF2BA02",
        7: "FFB8860B",
        8: "FFFFF2CC",
        12: "FF000000",
        13: "FF00B050",
        14: "FF3370FF",
        15: "FF000000",
        16: "FF000000",
    }
    for row, color in solid_rows.items():
        for column in range(1, 21):
            sheet.cell(row, column).font = Font(color=Color(rgb=color))
    for column in range(1, 21):
        sheet.cell(9, column).font = Font(
            color=Color(rgb="FFFFC000" if column <= 11 else "FFC00000")
        )
        sheet.cell(10, column).font = Font(
            color=Color(rgb="FFFFC000" if column <= 9 else "FFC00000")
        )
        sheet.cell(11, column).font = Font(
            color=Color(rgb="FFFFC000" if column <= 10 else "FFC00000")
        )
        if column >= 9:
            sheet.cell(15, column).font = Font(color=Color(rgb="FFFFC000"))
    sheet["G16"].font = Font(color=Color(rgb="FFED7D31"))
    workbook.save(source)
    workbook.close()

    result = aggregate_sales_workbooks(
        [SourceWorkbook("yellow-family", source)], template, output
    )

    assert result.signing_detail_count == 13
    assert result.signing_total == Decimal(91)
    workbook = load_workbook(output, data_only=False)
    sheet = workbook[SIGNING_TARGET]
    hidden_rows = {
        row for row in range(4, sheet.max_row + 1) if sheet.row_dimensions[row].hidden
    }
    assert hidden_rows == {4, 5, 6, 7, 8, 9, 10, 11, 16}
    assert sheet["C4"].value == "同一人"
    assert sheet["H4"].value == "明细1"
    assert sheet["I4"].value == 1
    assert _font_color_signature(sheet["C4"]) == ("rgb", "FFFFC000", 0.0)
    assert sheet.row_dimensions[15].hidden is False
    assert sheet["I17"].value == "=SUM(I4:I16)"
    assert all(not sheet.row_dimensions[row].hidden for row in range(17, 28))
    sheet.row_dimensions[4].hidden = False
    assert sheet["C4"].value == "同一人"
    assert sheet["I4"].value == 1
    workbook.close()


def test_each_c_to_g_key_field_can_trigger_hiding_but_other_columns_cannot(
    project_tmp_dir: Path,
) -> None:
    """C:G are the exact warm-color trigger boundary for one signing detail row."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "key-field-colors.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    _make_source(
        source,
        [
            _source_row(
                group="甲组" if index == 1 else None,
                sequence=index,
                person="同一人",
                customer=f"客户{index}",
                product=f"项目{index}",
                location=f"地点{index}",
                customer_source=f"来源{index}",
                months=(index,),
            )
            for index in range(1, 8)
        ],
        signing_sheet_name="签约数据",
    )

    workbook = load_workbook(source)
    sheet = workbook["签约数据"]
    for row in range(4, 11):
        for column in range(1, 21):
            sheet.cell(row, column).font = Font(color=Color(rgb="FF000000"))
    key_colors = ("FFFFC000", "FFFFC60A", "FFF2BA02", "FFED7D31", "FFF4B183")
    for row, (column, color) in enumerate(zip(range(3, 8), key_colors), start=4):
        sheet.cell(row, column).font = Font(color=Color(rgb=color))
    sheet["H9"].font = Font(color=Color(rgb="FFED7D31"))
    sheet["I10"].font = Font(color=Color(rgb="FFFFC000"))
    workbook.save(source)
    workbook.close()

    aggregate_sales_workbooks(
        [SourceWorkbook("key-field-colors", source)], template, output
    )

    workbook = load_workbook(output, data_only=False)
    sheet = workbook[SIGNING_TARGET]
    hidden_rows = {
        row for row in range(4, sheet.max_row + 1) if sheet.row_dimensions[row].hidden
    }
    assert hidden_rows == {4, 5, 6, 7, 8}
    assert sheet.row_dimensions[9].hidden is False
    assert sheet.row_dimensions[10].hidden is False
    workbook.close()


def test_yellow_style_on_blank_key_field_does_not_hide_black_text_row(
    project_tmp_dir: Path,
) -> None:
    """A blank C:G cell cannot hide a row whose populated key fields are black."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "blank-yellow-style.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    _make_source(
        source,
        [
            _source_row(
                group="甲组",
                sequence=1,
                person="黑色人员",
                customer="黑色客户",
                product="黑色项目",
                location="黑色地点",
                customer_source=None,
                months=(10,),
            )
        ],
        signing_sheet_name="签约数据",
    )

    workbook = load_workbook(source)
    sheet = workbook["签约数据"]
    for column in range(3, 7):
        sheet.cell(4, column).font = Font(color=Color(rgb="FF000000"))
    sheet["G4"].font = Font(color=Color(rgb="FFFFC000"))
    workbook.save(source)
    workbook.close()

    aggregate_sales_workbooks(
        [SourceWorkbook("blank-yellow-style", source)], template, output
    )

    workbook = load_workbook(output, data_only=False)
    sheet = workbook[SIGNING_TARGET]
    assert sheet.row_dimensions[4].hidden is False
    assert sheet["C4"].font.color.rgb == "FF000000"
    assert sheet["G4"].value is None
    assert sheet["G4"].font.color.rgb == "FFFFC000"
    assert sheet["I4"].value == 10
    workbook.close()


def test_theme_color_indexes_use_dark_then_light_ooxml_order(
    project_tmp_dir: Path,
) -> None:
    """Theme 0 remains dark even when theme 1 is customized to yellow."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "theme-order.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    _make_source(
        source,
        [
            _source_row(group="甲组", sequence=1, person="同一人", months=(1,)),
            _source_row(sequence=2, person="同一人", months=(2,)),
        ],
        signing_sheet_name="签约情况",
    )

    workbook = load_workbook(source)
    workbook.loaded_theme = theme_xml.replace(
        'lastClr="FFFFFF"', 'lastClr="FFC000"', 1
    ).encode("utf-8")
    sheet = workbook["签约情况"]
    sheet["C4"].font = Font(color=Color(theme=0))
    sheet["C5"].font = Font(color=Color(theme=1))
    workbook.save(source)
    workbook.close()

    aggregate_sales_workbooks(
        [SourceWorkbook("theme-order", source)], template, output
    )

    workbook = load_workbook(output, data_only=False)
    sheet = workbook[SIGNING_TARGET]
    assert sheet.row_dimensions[4].hidden is False
    assert sheet.row_dimensions[5].hidden is True
    workbook.close()


def test_theme_indexed_and_tinted_yellow_fonts_are_hidden(
    project_tmp_dir: Path,
) -> None:
    """Static theme/indexed yellow resolves while auto, missing, blue, and red stay visible."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "encoded-colors.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    _make_source(
        source,
        [
            _source_row(
                group="甲组" if index == 1 else None,
                sequence=index,
                person="同一人",
                months=(index,),
            )
            for index in range(1, 8)
        ],
        signing_sheet_name="签约情况",
    )

    workbook = load_workbook(source)
    workbook.loaded_theme = theme_xml.replace(
        'val="8064A2"', 'val="FFC000"'
    ).encode("utf-8")
    sheet = workbook["签约情况"]
    encoded_colors = {
        4: Color(theme=7, tint=0.25),
        5: Color(indexed=5),
        6: Color(theme=7, tint=-0.25),
        7: Color(theme=4),
        8: Color(indexed=2),
        9: Color(auto=True),
        10: None,
    }
    for row, color in encoded_colors.items():
        for column in range(1, 21):
            sheet.cell(row, column).font = Font(color=color)
    workbook.save(source)
    workbook.close()

    aggregate_sales_workbooks(
        [SourceWorkbook("encoded-colors", source)], template, output
    )

    workbook = load_workbook(output, data_only=False)
    sheet = workbook[SIGNING_TARGET]
    hidden_rows = {
        row for row in range(4, sheet.max_row + 1) if sheet.row_dimensions[row].hidden
    }
    assert hidden_rows == {4, 5, 6}
    assert all(not sheet.row_dimensions[row].hidden for row in range(7, sheet.max_row + 1))
    workbook.close()


def test_inherited_group_keeps_group_header_font_color(project_tmp_dir: Path) -> None:
    """A blank detail A cell inherits both the last group value and its font color."""

    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "inherited-group-color.xlsx"
    output = project_tmp_dir / "output.xlsx"
    _make_template(template)
    _make_source(
        source,
        [
            _source_row(group="甲组"),
            _source_row(sequence=1, person="A", months=(10,)),
        ],
        signing_sheet_name="签约情况",
    )
    workbook = load_workbook(source)
    sheet = workbook["签约情况"]
    sheet["A4"].font = Font(color=Color(rgb="FF00B050"))
    sheet["A5"].font = Font(color=Color(rgb="FFFF0000"))
    workbook.save(source)
    workbook.close()

    aggregate_sales_workbooks([SourceWorkbook("source", source)], template, output)

    workbook = load_workbook(output, data_only=False)
    assert _font_color_signature(workbook[SIGNING_TARGET]["A4"]) == (
        "rgb",
        "FF00B050",
        0.0,
    )
    workbook.close()


def test_repository_template_with_shared_strings_reopens(
    project_tmp_dir: Path,
) -> None:
    """Preserved non-target worksheets keep their shared-string package dependency."""

    template = (
        Path(__file__).resolve().parents[1]
        / "excel_file_example"
        / "汇总效果-合并版-2026年销售数据统计2.xlsx"
    )
    source = project_tmp_dir / "shared-strings-source.xlsx"
    output = project_tmp_dir / "shared-strings-output.xlsx"
    _make_source(
        source,
        [_source_row(group="甲组", sequence=1, person="A", months=(10,))],
        signing_sheet_name="签约情况",
    )

    aggregate_sales_workbooks([SourceWorkbook("source", source)], template, output)

    workbook = load_workbook(output, data_only=False)
    assert workbook[SIGNING_TARGET]["C4"].value == "A"
    workbook.close()


def test_empty_signing_data_is_rejected_during_aggregation(
    project_tmp_dir: Path,
) -> None:
    template = project_tmp_dir / "template.xlsx"
    source = project_tmp_dir / "empty.xlsx"
    _make_template(template)
    _make_source(
        source,
        [
            _source_row(sequence="个人月度合计：", months=(" - ",)),
            _source_row(sequence="仅供说明", note="没有业务明细"),
        ],
        signing_sheet_name="Sheet1",
    )

    with pytest.raises(SalesAggregationError, match="没有有效签约明细"):
        aggregate_sales_workbooks(
            [SourceWorkbook("empty", source)],
            template,
            project_tmp_dir / "out.xlsx",
        )
