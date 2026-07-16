"""Rebuild sales signing and repayment summaries from ordered XLSX sources."""

from __future__ import annotations

from collections import OrderedDict
from copy import copy
from dataclasses import dataclass
from decimal import Decimal
import os
from pathlib import Path, PurePosixPath
import posixpath
import re
import unicodedata
from typing import Any, Iterable, Sequence
from uuid import uuid4
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


SIGNING_SOURCE_SHEET = "Sheet1"
REPAYMENT_SOURCE_SHEET = "Sheet2"
SIGNING_TARGET_SHEET = "展示用-2026年签约数据汇总（0702统计）"
REPAYMENT_TARGET_SHEET = "展示用-2026年回款数据汇总(0702统计)"

SIGNING_PERSON_MONTH = "个人月度合计："
SIGNING_PERSON_QUARTER = "个人季度合计："
SIGNING_PERSON_YEAR = "个人年度合计："
SIGNING_GROUP_MONTH = "小组月度合计："
SIGNING_GROUP_QUARTER = "小组季度合计："
SIGNING_GROUP_YEAR = "小组年度合计："
SIGNING_DEPT_MONTH = "部门月度合计："
SIGNING_DEPT_QUARTER = "部门季度合计："
SIGNING_DEPT_YEAR = "部门年度合计："
REPAYMENT_PERSON_TOTAL = "个人小计："
REPAYMENT_DEPT_TOTAL = "部门总计："

FORMULA_ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
XML_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XML_DOC_REL_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
XML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


class SalesAggregationError(RuntimeError):
    """Base error for a rejected or failed aggregation task."""


class DuplicateSourceError(SalesAggregationError):
    """Raised when one stable source identifier appears twice in a task."""


class SourceValidationError(SalesAggregationError):
    """Raised when an input workbook does not match the confirmed SOP."""


class TemplateValidationError(SalesAggregationError):
    """Raised when the target workbook cannot supply the required structure."""


@dataclass(frozen=True, slots=True)
class SourceWorkbook:
    """One ordered source workbook and its stable task-level identity."""

    source_file_id: str
    path: Path

    def __init__(self, source_file_id: str, path: str | Path) -> None:
        object.__setattr__(self, "source_file_id", source_file_id.strip())
        object.__setattr__(self, "path", Path(path).resolve())


@dataclass(frozen=True, slots=True)
class AggregationResult:
    """Control totals and the generated XLSX path."""

    output_path: Path
    source_count: int
    signing_detail_count: int
    repayment_detail_count: int
    signing_total: Decimal
    repayment_current_year_total: Decimal
    repayment_contract_total: Decimal
    repayment_cumulative_total: Decimal


@dataclass(frozen=True, slots=True)
class SigningRecord:
    source_file_id: str
    source_path: Path
    source_row: int
    group: str
    person: str
    values: tuple[Any, ...]
    months: tuple[Any, ...]


@dataclass(frozen=True, slots=True)
class RepaymentRecord:
    source_file_id: str
    source_path: Path
    source_row: int
    person: str
    values: tuple[Any, ...]
    months: tuple[Any, ...]


@dataclass(frozen=True, slots=True)
class ParsedSource:
    signing: tuple[SigningRecord, ...]
    repayment: tuple[RepaymentRecord, ...]


@dataclass(frozen=True, slots=True)
class RowStyleSample:
    styles: tuple[Any, ...]
    height: float | None


@dataclass(frozen=True, slots=True)
class SigningStyleSamples:
    detail: RowStyleSample
    personal_month: RowStyleSample
    personal_quarter: RowStyleSample
    personal_year: RowStyleSample
    group_month: RowStyleSample
    group_quarter: RowStyleSample
    group_year: RowStyleSample
    department_month: RowStyleSample
    department_quarter: RowStyleSample
    department_year: RowStyleSample
    person_spacer: RowStyleSample
    group_spacer: RowStyleSample


@dataclass(frozen=True, slots=True)
class RepaymentStyleSamples:
    detail: RowStyleSample
    person_total: RowStyleSample
    spacer: RowStyleSample
    department_total: RowStyleSample


def _normalized_header(value: Any) -> str:
    if value is None:
        return ""
    normalized = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", "", normalized)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def _as_decimal(value: Any) -> Decimal:
    if _is_blank(value) or value == "/":
        return Decimal(0)
    if not _is_numeric(value):
        raise TypeError(f"not numeric: {value!r}")
    return Decimal(str(value))


def _source_error(
    source: SourceWorkbook,
    message: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
    column: str | None = None,
) -> SourceValidationError:
    location = [source.path.name]
    if sheet:
        location.append(sheet)
    if row is not None:
        location.append(f"第 {row} 行")
    if column:
        location.append(f"{column} 列")
    return SourceValidationError(f"{' / '.join(location)}：{message}")


def _validate_source_identity(source: SourceWorkbook) -> None:
    if not source.source_file_id:
        raise SourceValidationError(f"{source.path.name}：source_file_id 不能为空")
    if source.path.suffix.lower() != ".xlsx":
        raise SourceValidationError(f"{source.path.name}：销售汇总只支持 .xlsx 文件")
    if not source.path.is_file():
        raise SourceValidationError(f"{source.path.name}：源文件不存在")


def _validate_headers(source: SourceWorkbook, workbook: Any) -> None:
    missing = [
        name
        for name in (SIGNING_SOURCE_SHEET, REPAYMENT_SOURCE_SHEET)
        if name not in workbook.sheetnames
    ]
    if missing:
        raise _source_error(source, f"缺少工作表：{'、'.join(missing)}")

    signing = workbook[SIGNING_SOURCE_SHEET]
    signing_row2 = {
        "A": "组别",
        "B": "序号",
        "C": "人员",
        "D": "客户简称",
        "E": "销售产品名称/项目名称",
        "F": "落地地点",
        "G": "客户来源（展会需注明）",
        "H": "备注",
        "I": "签约后且收到第一笔款后才算已签约",
    }
    for column, expected in signing_row2.items():
        actual = signing[f"{column}2"].value
        if _normalized_header(actual) != _normalized_header(expected):
            raise _source_error(
                source,
                f"表头不匹配，期望“{expected}”",
                sheet=SIGNING_SOURCE_SHEET,
                row=2,
                column=column,
            )
    for index in range(12):
        column = get_column_letter(9 + index)
        expected = f"{index + 1}月"
        if _normalized_header(signing[f"{column}3"].value) != expected:
            raise _source_error(
                source,
                f"月份表头不匹配，期望“{expected}”",
                sheet=SIGNING_SOURCE_SHEET,
                row=3,
                column=column,
            )

    repayment = workbook[REPAYMENT_SOURCE_SHEET]
    repayment_row2 = {
        "A": "姓名",
        "B": "收款主体",
        "C": "付款主体",
        "D": "产品",
        "E": "合同日期",
        "F": "合同总价",
        "G": "居间费",
        "H": "已回款-提成口（26年以前）",
        "I": "已回款-提成口径（26年）",
        "J": "已回款-提成口径总额",
        "K": "待回款额",
        "L": "已开发票金额",
        "M": "已开票未回款",
        "N": "回款比例",
        "O": "税/运费",
        "P": "预计剩余回款时间及金额",
        "Q": "备注",
        "R": "已回款-业绩口径(26年)",
    }
    for column, expected in repayment_row2.items():
        actual = repayment[f"{column}2"].value
        if _normalized_header(actual) != _normalized_header(expected):
            raise _source_error(
                source,
                f"表头不匹配，期望“{expected}”",
                sheet=REPAYMENT_SOURCE_SHEET,
                row=2,
                column=column,
            )
    for index in range(12):
        column = get_column_letter(18 + index)
        expected = f"{index + 1}月"
        if _normalized_header(repayment[f"{column}3"].value) != expected:
            raise _source_error(
                source,
                f"月份表头不匹配，期望“{expected}”",
                sheet=REPAYMENT_SOURCE_SHEET,
                row=3,
                column=column,
            )


def _parse_signing(source: SourceWorkbook, sheet: Worksheet) -> tuple[SigningRecord, ...]:
    records: list[SigningRecord] = []
    current_group: str | None = None
    for row in range(4, sheet.max_row + 1):
        values = tuple(sheet.cell(row, column).value for column in range(1, 21))
        group_value = values[0]
        if not _is_blank(group_value):
            if not isinstance(group_value, str):
                raise _source_error(
                    source,
                    "组别必须是文本",
                    sheet=SIGNING_SOURCE_SHEET,
                    row=row,
                    column="A",
                )
            current_group = group_value

        if all(_is_blank(value) for value in values[1:]):
            continue
        if current_group is None:
            raise _source_error(
                source,
                "存在业务内容但无法取得组别",
                sheet=SIGNING_SOURCE_SHEET,
                row=row,
                column="A",
            )
        person = values[2]
        if _is_blank(person) or not isinstance(person, str):
            raise _source_error(
                source,
                "人员不能为空且必须是文本",
                sheet=SIGNING_SOURCE_SHEET,
                row=row,
                column="C",
            )
        months = values[8:20]
        for offset, value in enumerate(months, start=9):
            if not _is_blank(value) and not _is_numeric(value):
                raise _source_error(
                    source,
                    "签约金额必须是数值或空值",
                    sheet=SIGNING_SOURCE_SHEET,
                    row=row,
                    column=get_column_letter(offset),
                )
        output_values = (current_group, *values[1:])
        records.append(
            SigningRecord(
                source_file_id=source.source_file_id,
                source_path=source.path,
                source_row=row,
                group=current_group,
                person=person,
                values=tuple(output_values),
                months=tuple(months),
            )
        )
    return tuple(records)


def _parse_repayment(
    source: SourceWorkbook, sheet: Worksheet
) -> tuple[RepaymentRecord, ...]:
    records: list[RepaymentRecord] = []
    numeric_columns = (6, 7, 8, 12, 13, 15, *range(18, 30))
    for row in range(4, sheet.max_row + 1):
        values = tuple(sheet.cell(row, column).value for column in range(1, 30))
        if all(_is_blank(value) for value in values):
            continue
        person = values[0]
        if _is_blank(person) or not isinstance(person, str):
            raise _source_error(
                source,
                "姓名不能为空且必须是文本",
                sheet=REPAYMENT_SOURCE_SHEET,
                row=row,
                column="A",
            )
        for column in numeric_columns:
            value = values[column - 1]
            if _is_blank(value) or value == "/" or _is_numeric(value):
                continue
            raise _source_error(
                source,
                "金额必须是数值、空值或 /",
                sheet=REPAYMENT_SOURCE_SHEET,
                row=row,
                column=get_column_letter(column),
            )
        records.append(
            RepaymentRecord(
                source_file_id=source.source_file_id,
                source_path=source.path,
                source_row=row,
                person=person,
                values=values,
                months=tuple(values[17:29]),
            )
        )
    return tuple(records)


def _parse_source(source: SourceWorkbook) -> ParsedSource:
    _validate_source_identity(source)
    try:
        workbook = load_workbook(source.path, data_only=False, read_only=False)
    except Exception as exc:
        raise _source_error(source, f"无法打开 XLSX：{exc}") from exc
    try:
        _validate_headers(source, workbook)
        return ParsedSource(
            signing=_parse_signing(source, workbook[SIGNING_SOURCE_SHEET]),
            repayment=_parse_repayment(source, workbook[REPAYMENT_SOURCE_SHEET]),
        )
    finally:
        workbook.close()


def validate_source_workbook(source: SourceWorkbook) -> tuple[int, int]:
    """Validate one source and return its signing and repayment detail counts."""

    parsed = _parse_source(source)
    return len(parsed.signing), len(parsed.repayment)


def _capture_row_style(sheet: Worksheet, row: int, max_column: int) -> RowStyleSample:
    return RowStyleSample(
        styles=tuple(copy(sheet.cell(row, column)._style) for column in range(1, max_column + 1)),
        height=sheet.row_dimensions[row].height,
    )


def _apply_row_style(
    sheet: Worksheet, row: int, sample: RowStyleSample, max_column: int
) -> None:
    for column in range(1, max_column + 1):
        sheet.cell(row, column)._style = copy(sample.styles[column - 1])
    dimension = sheet.row_dimensions[row]
    dimension.height = sample.height
    dimension.hidden = False
    dimension.collapsed = False
    dimension.outlineLevel = 0


def _find_label_row(sheet: Worksheet, column: int, label: str) -> int:
    expected = _normalized_header(label)
    for row in range(4, sheet.max_row + 1):
        if _normalized_header(sheet.cell(row, column).value) == expected:
            return row
    raise TemplateValidationError(f"{sheet.title}：缺少格式样本“{label}”")


def _find_detail_row(sheet: Worksheet, person_column: int, max_column: int) -> int:
    labels = {
        _normalized_header(label)
        for label in (
            SIGNING_PERSON_MONTH,
            SIGNING_PERSON_QUARTER,
            SIGNING_PERSON_YEAR,
            SIGNING_GROUP_MONTH,
            SIGNING_GROUP_QUARTER,
            SIGNING_GROUP_YEAR,
            SIGNING_DEPT_MONTH,
            SIGNING_DEPT_QUARTER,
            SIGNING_DEPT_YEAR,
            REPAYMENT_PERSON_TOTAL,
            REPAYMENT_DEPT_TOTAL,
        )
    }
    for row in range(4, sheet.max_row + 1):
        person = sheet.cell(row, person_column).value
        if _is_blank(person):
            continue
        row_labels = {
            _normalized_header(sheet.cell(row, column).value)
            for column in range(1, min(max_column, 8) + 1)
        }
        if not row_labels.intersection(labels):
            return row
    raise TemplateValidationError(f"{sheet.title}：缺少明细行格式样本")


def _find_blank_row_after(sheet: Worksheet, row: int, max_column: int) -> int:
    candidate = row + 1
    if candidate <= sheet.max_row and all(
        _is_blank(sheet.cell(candidate, column).value)
        for column in range(1, max_column + 1)
    ):
        return candidate
    for candidate in range(4, sheet.max_row + 1):
        if all(
            _is_blank(sheet.cell(candidate, column).value)
            for column in range(1, max_column + 1)
        ):
            return candidate
    raise TemplateValidationError(f"{sheet.title}：缺少空行格式样本")


def _capture_signing_samples(sheet: Worksheet) -> SigningStyleSamples:
    rows = {
        "detail": _find_detail_row(sheet, 3, 20),
        "personal_month": _find_label_row(sheet, 2, SIGNING_PERSON_MONTH),
        "personal_quarter": _find_label_row(sheet, 2, SIGNING_PERSON_QUARTER),
        "personal_year": _find_label_row(sheet, 2, SIGNING_PERSON_YEAR),
        "group_month": _find_label_row(sheet, 2, SIGNING_GROUP_MONTH),
        "group_quarter": _find_label_row(sheet, 2, SIGNING_GROUP_QUARTER),
        "group_year": _find_label_row(sheet, 2, SIGNING_GROUP_YEAR),
        "department_month": _find_label_row(sheet, 1, SIGNING_DEPT_MONTH),
        "department_quarter": _find_label_row(sheet, 1, SIGNING_DEPT_QUARTER),
        "department_year": _find_label_row(sheet, 1, SIGNING_DEPT_YEAR),
    }
    person_spacer = _find_blank_row_after(sheet, rows["personal_year"], 20)
    group_spacer = _find_blank_row_after(sheet, rows["group_year"], 20)
    return SigningStyleSamples(
        **{
            key: _capture_row_style(sheet, row, 20)
            for key, row in rows.items()
        },
        person_spacer=_capture_row_style(sheet, person_spacer, 20),
        group_spacer=_capture_row_style(sheet, group_spacer, 20),
    )


def _capture_repayment_samples(sheet: Worksheet) -> RepaymentStyleSamples:
    detail = _find_detail_row(sheet, 1, 29)
    person_total = _find_label_row(sheet, 4, REPAYMENT_PERSON_TOTAL)
    department_total = _find_label_row(sheet, 4, REPAYMENT_DEPT_TOTAL)
    spacer = _find_blank_row_after(sheet, person_total, 29)
    return RepaymentStyleSamples(
        detail=_capture_row_style(sheet, detail, 29),
        person_total=_capture_row_style(sheet, person_total, 29),
        spacer=_capture_row_style(sheet, spacer, 29),
        department_total=_capture_row_style(sheet, department_total, 29),
    )


def _clear_business_area(sheet: Worksheet) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row >= 4:
            sheet.unmerge_cells(str(merged_range))
    if sheet.max_row >= 4:
        sheet.delete_rows(4, sheet.max_row - 3)
    for index in list(sheet.row_dimensions):
        if isinstance(index, int) and index >= 4:
            del sheet.row_dimensions[index]


def _merge_label(sheet: Worksheet, row: int, start_column: int = 2) -> None:
    sheet.merge_cells(
        start_row=row,
        start_column=start_column,
        end_row=row,
        end_column=8,
    )


def _write_month_quarter_year(
    sheet: Worksheet,
    month_row: int,
    quarter_row: int,
    year_row: int,
    month_formula_rows: Sequence[int] | tuple[int, int],
    labels: tuple[str, str, str],
) -> None:
    _merge_label(sheet, month_row)
    _merge_label(sheet, quarter_row)
    _merge_label(sheet, year_row)
    sheet.cell(month_row, 2).value = labels[0]
    sheet.cell(quarter_row, 2).value = labels[1]
    sheet.cell(year_row, 2).value = labels[2]

    if isinstance(month_formula_rows, tuple):
        start, end = month_formula_rows
        for column in range(9, 21):
            letter = get_column_letter(column)
            sheet.cell(month_row, column).value = f"=SUM({letter}{start}:{letter}{end})"
    else:
        references = list(month_formula_rows)
        for column in range(9, 21):
            letter = get_column_letter(column)
            args = ",".join(f"{letter}{row}" for row in references)
            sheet.cell(month_row, column).value = f"=SUM({args})"

    quarter_ranges = ((9, 11), (12, 14), (15, 17), (18, 20))
    for start_column, end_column in quarter_ranges:
        sheet.merge_cells(
            start_row=quarter_row,
            start_column=start_column,
            end_row=quarter_row,
            end_column=end_column,
        )
        start_letter = get_column_letter(start_column)
        end_letter = get_column_letter(end_column)
        sheet.cell(quarter_row, start_column).value = (
            f"=SUM({start_letter}{month_row}:{end_letter}{month_row})"
        )
    sheet.merge_cells(
        start_row=year_row,
        start_column=9,
        end_row=year_row,
        end_column=20,
    )
    sheet.cell(year_row, 9).value = f"=SUM(I{month_row}:T{month_row})"


def _write_signing_sheet(
    sheet: Worksheet,
    records: Sequence[SigningRecord],
    samples: SigningStyleSamples,
) -> int:
    grouped: OrderedDict[str, OrderedDict[str, list[SigningRecord]]] = OrderedDict()
    for record in records:
        grouped.setdefault(record.group, OrderedDict()).setdefault(record.person, []).append(record)

    row = 4
    group_month_rows: list[int] = []
    groups = list(grouped.items())
    for group_index, (group, people) in enumerate(groups):
        group_start = row
        person_month_rows: list[int] = []
        people_items = list(people.items())
        for person_index, (_, person_records) in enumerate(people_items):
            detail_start = row
            for record in person_records:
                _apply_row_style(sheet, row, samples.detail, 20)
                for column, value in enumerate(record.values, start=1):
                    sheet.cell(row, column).value = value
                row += 1
            detail_end = row - 1

            personal_month = row
            _apply_row_style(sheet, personal_month, samples.personal_month, 20)
            _apply_row_style(sheet, personal_month + 1, samples.personal_quarter, 20)
            _apply_row_style(sheet, personal_month + 2, samples.personal_year, 20)
            _write_month_quarter_year(
                sheet,
                personal_month,
                personal_month + 1,
                personal_month + 2,
                (detail_start, detail_end),
                (SIGNING_PERSON_MONTH, SIGNING_PERSON_QUARTER, SIGNING_PERSON_YEAR),
            )
            person_month_rows.append(personal_month)
            row += 3
            if person_index < len(people_items) - 1:
                _apply_row_style(sheet, row, samples.person_spacer, 20)
                row += 1

        group_month = row
        _apply_row_style(sheet, group_month, samples.group_month, 20)
        _apply_row_style(sheet, group_month + 1, samples.group_quarter, 20)
        _apply_row_style(sheet, group_month + 2, samples.group_year, 20)
        _write_month_quarter_year(
            sheet,
            group_month,
            group_month + 1,
            group_month + 2,
            person_month_rows,
            (SIGNING_GROUP_MONTH, SIGNING_GROUP_QUARTER, SIGNING_GROUP_YEAR),
        )
        group_month_rows.append(group_month)
        group_end = group_month + 2
        sheet.merge_cells(
            start_row=group_start,
            start_column=1,
            end_row=group_end,
            end_column=1,
        )
        sheet.cell(group_start, 1).value = group
        row += 3
        if group_index < len(groups) - 1:
            _apply_row_style(sheet, row, samples.group_spacer, 20)
            row += 1

    department_month = row
    _apply_row_style(sheet, department_month, samples.department_month, 20)
    _apply_row_style(sheet, department_month + 1, samples.department_quarter, 20)
    _apply_row_style(sheet, department_month + 2, samples.department_year, 20)
    _write_month_quarter_year(
        sheet,
        department_month,
        department_month + 1,
        department_month + 2,
        group_month_rows,
        (SIGNING_DEPT_MONTH, SIGNING_DEPT_QUARTER, SIGNING_DEPT_YEAR),
    )
    for target_row in range(department_month, department_month + 3):
        sheet.unmerge_cells(start_row=target_row, start_column=2, end_row=target_row, end_column=8)
        sheet.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=8)
        sheet.cell(target_row, 1).value = (
            SIGNING_DEPT_MONTH,
            SIGNING_DEPT_QUARTER,
            SIGNING_DEPT_YEAR,
        )[target_row - department_month]
    return department_month + 2


def _write_repayment_sheet(
    sheet: Worksheet,
    records: Sequence[RepaymentRecord],
    samples: RepaymentStyleSamples,
) -> int:
    grouped: OrderedDict[str, list[RepaymentRecord]] = OrderedDict()
    for record in records:
        grouped.setdefault(record.person, []).append(record)

    row = 4
    subtotal_rows: list[int] = []
    people = list(grouped.items())
    for person_index, (_, person_records) in enumerate(people):
        detail_start = row
        for record in person_records:
            _apply_row_style(sheet, row, samples.detail, 29)
            direct_columns = (*range(1, 9), 12, 13, *range(15, 30))
            for column in direct_columns:
                sheet.cell(row, column).value = record.values[column - 1]
            sheet.cell(row, 9).value = f"=SUM(R{row}:AC{row})"
            sheet.cell(row, 10).value = f"=SUM(H{row}:I{row})"
            sheet.cell(row, 11).value = f"=F{row}-J{row}"
            sheet.cell(row, 14).value = f"=IFERROR(J{row}/F{row},0)"
            row += 1
        detail_end = row - 1

        subtotal = row
        _apply_row_style(sheet, subtotal, samples.person_total, 29)
        sheet.cell(subtotal, 4).value = REPAYMENT_PERSON_TOTAL
        for column in (*range(6, 14), 15, *range(18, 30)):
            letter = get_column_letter(column)
            sheet.cell(subtotal, column).value = (
                f"=SUM({letter}{detail_start}:{letter}{detail_end})"
            )
        sheet.cell(subtotal, 14).value = f"=IFERROR(J{subtotal}/F{subtotal},0)"
        subtotal_rows.append(subtotal)
        row += 1
        if person_index < len(people) - 1:
            _apply_row_style(sheet, row, samples.spacer, 29)
            row += 1

    department = row
    _apply_row_style(sheet, department, samples.department_total, 29)
    sheet.cell(department, 4).value = REPAYMENT_DEPT_TOTAL
    for column in (*range(6, 14), 15, *range(18, 30)):
        letter = get_column_letter(column)
        references = ",".join(f"{letter}{subtotal}" for subtotal in subtotal_rows)
        sheet.cell(department, column).value = f"=SUM({references})"
    sheet.cell(department, 14).value = f"=IFERROR(J{department}/F{department},0)"
    return department


def _validate_template(workbook: Any) -> tuple[Worksheet, Worksheet]:
    missing = [
        name
        for name in (SIGNING_TARGET_SHEET, REPAYMENT_TARGET_SHEET)
        if name not in workbook.sheetnames
    ]
    if missing:
        raise TemplateValidationError(f"目标模板缺少工作表：{'、'.join(missing)}")
    return workbook[SIGNING_TARGET_SHEET], workbook[REPAYMENT_TARGET_SHEET]


def _formula_scan(sheet: Worksheet, max_row: int, max_column: int) -> None:
    for row in sheet.iter_rows(
        min_row=4, max_row=max_row, min_col=1, max_col=max_column
    ):
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            upper = value.upper()
            if any(token in upper for token in FORMULA_ERROR_TOKENS):
                raise SalesAggregationError(
                    f"{sheet.title} / {cell.coordinate}：发现公式错误 {value}"
                )


def _worksheet_parts_by_name(path: Path) -> dict[str, str]:
    with zipfile.ZipFile(path) as archive:
        workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rels = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_xml.findall(f"{{{XML_REL_NS}}}Relationship")
    }
    parts: dict[str, str] = {}
    sheets = workbook_xml.find(f"{{{XML_MAIN_NS}}}sheets")
    if sheets is None:
        return parts
    for sheet in sheets:
        rel_id = sheet.attrib[f"{{{XML_DOC_REL_NS}}}id"]
        target = rels[rel_id]
        if target.startswith("/"):
            part = target.lstrip("/")
        else:
            part = posixpath.normpath(posixpath.join("xl", target))
        parts[sheet.attrib["name"]] = part
    return parts


def _rels_part_name(part_name: str) -> str:
    part = PurePosixPath(part_name)
    return str(part.parent / "_rels" / f"{part.name}.rels")


def _collect_related_parts(archive: zipfile.ZipFile, root_parts: Iterable[str]) -> set[str]:
    existing = set(archive.namelist())
    collected: set[str] = set()
    pending = list(root_parts)
    while pending:
        part = pending.pop()
        if part in collected or part not in existing:
            continue
        collected.add(part)
        rels_name = _rels_part_name(part)
        if rels_name not in existing:
            continue
        collected.add(rels_name)
        rels_xml = ET.fromstring(archive.read(rels_name))
        for rel in rels_xml.findall(f"{{{XML_REL_NS}}}Relationship"):
            if rel.attrib.get("TargetMode") == "External":
                continue
            target = rel.attrib.get("Target", "")
            if target.startswith("/"):
                target_part = target.lstrip("/")
            else:
                target_part = posixpath.normpath(
                    posixpath.join(posixpath.dirname(part), target)
                )
            if target_part in existing and target_part not in collected:
                pending.append(target_part)
    return collected


def _restore_non_target_parts(
    template_path: Path, generated_path: Path, patched_path: Path
) -> None:
    sheet_parts = _worksheet_parts_by_name(template_path)
    non_target_roots = [
        part
        for name, part in sheet_parts.items()
        if name not in {SIGNING_TARGET_SHEET, REPAYMENT_TARGET_SHEET}
    ]
    with zipfile.ZipFile(template_path) as source_zip:
        preserved = _collect_related_parts(source_zip, non_target_roots)
        preserved.add("[Content_Types].xml")
        source_payload = {name: source_zip.read(name) for name in preserved}
    with zipfile.ZipFile(generated_path) as generated_zip, zipfile.ZipFile(
        patched_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as output_zip:
        for entry in generated_zip.infolist():
            if entry.filename in preserved:
                continue
            output_zip.writestr(entry, generated_zip.read(entry.filename))
        for name, payload in source_payload.items():
            output_zip.writestr(name, payload)


def _verify_output(
    output_path: Path,
    *,
    signing_last_row: int,
    repayment_last_row: int,
) -> None:
    workbook = load_workbook(output_path, data_only=False, read_only=False)
    try:
        signing, repayment = _validate_template(workbook)
        _formula_scan(signing, signing_last_row, 20)
        _formula_scan(repayment, repayment_last_row, 29)
        for sheet, last_row in (
            (signing, signing_last_row),
            (repayment, repayment_last_row),
        ):
            for row in range(4, last_row + 1):
                if sheet.row_dimensions[row].hidden:
                    raise SalesAggregationError(
                        f"{sheet.title} / 第 {row} 行：生成结果中存在隐藏行"
                    )
    finally:
        workbook.close()


def _validate_source_list(sources: Sequence[SourceWorkbook]) -> None:
    if not sources:
        raise SalesAggregationError("至少需要一个源文件")
    seen: set[str] = set()
    for source in sources:
        if source.source_file_id in seen:
            raise DuplicateSourceError(
                f"同一任务重复包含源文件：{source.source_file_id}"
            )
        seen.add(source.source_file_id)


def aggregate_sales_workbooks(
    sources: Sequence[SourceWorkbook],
    template_path: str | Path,
    output_path: str | Path,
) -> AggregationResult:
    """Create one XLSX summary from ordered sources and a formatted template."""

    sources = tuple(sources)
    _validate_source_list(sources)
    template = Path(template_path).resolve()
    output = Path(output_path).resolve()
    if not template.is_file() or template.suffix.lower() != ".xlsx":
        raise TemplateValidationError("目标模板不存在或不是 .xlsx 文件")
    protected_paths = {template, *(source.path for source in sources)}
    if output in protected_paths:
        raise SalesAggregationError("输出文件不得覆盖模板或任一源文件")

    signing_records: list[SigningRecord] = []
    repayment_records: list[RepaymentRecord] = []
    for source in sources:
        parsed = _parse_source(source)
        signing_records.extend(parsed.signing)
        repayment_records.extend(parsed.repayment)
    if not signing_records:
        raise SalesAggregationError("所有源文件均没有有效签约明细")
    if not repayment_records:
        raise SalesAggregationError("所有源文件均没有有效回款明细")

    signing_total = sum(
        (_as_decimal(value) for record in signing_records for value in record.months),
        Decimal(0),
    )
    repayment_current_year_total = sum(
        (_as_decimal(value) for record in repayment_records for value in record.months),
        Decimal(0),
    )
    repayment_contract_total = sum(
        (_as_decimal(record.values[5]) for record in repayment_records), Decimal(0)
    )
    repayment_prior_total = sum(
        (_as_decimal(record.values[7]) for record in repayment_records), Decimal(0)
    )
    repayment_cumulative_total = repayment_prior_total + repayment_current_year_total

    try:
        workbook = load_workbook(template, data_only=False, read_only=False)
    except Exception as exc:
        raise TemplateValidationError(f"无法打开目标模板：{exc}") from exc
    try:
        signing_sheet, repayment_sheet = _validate_template(workbook)
        signing_samples = _capture_signing_samples(signing_sheet)
        repayment_samples = _capture_repayment_samples(repayment_sheet)
        _clear_business_area(signing_sheet)
        _clear_business_area(repayment_sheet)
        signing_last_row = _write_signing_sheet(
            signing_sheet, signing_records, signing_samples
        )
        repayment_last_row = _write_repayment_sheet(
            repayment_sheet, repayment_records, repayment_samples
        )
        _formula_scan(signing_sheet, signing_last_row, 20)
        _formula_scan(repayment_sheet, repayment_last_row, 29)
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

        output.parent.mkdir(parents=True, exist_ok=True)
        raw_temp = output.parent / f".{output.name}.{uuid4().hex}.raw.xlsx"
        patched_temp = output.parent / f".{output.name}.{uuid4().hex}.tmp.xlsx"
        try:
            workbook.save(raw_temp)
            _restore_non_target_parts(template, raw_temp, patched_temp)
            _verify_output(
                patched_temp,
                signing_last_row=signing_last_row,
                repayment_last_row=repayment_last_row,
            )
            os.replace(patched_temp, output)
        finally:
            raw_temp.unlink(missing_ok=True)
            patched_temp.unlink(missing_ok=True)
    finally:
        workbook.close()

    return AggregationResult(
        output_path=output,
        source_count=len(sources),
        signing_detail_count=len(signing_records),
        repayment_detail_count=len(repayment_records),
        signing_total=signing_total,
        repayment_current_year_total=repayment_current_year_total,
        repayment_contract_total=repayment_contract_total,
        repayment_cumulative_total=repayment_cumulative_total,
    )
